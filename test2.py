# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# import os
#
#
# serv_obj =Service("")
# driver=webdriver.Chrome(service=serv_obj)
# driver.get("https://opensource-demo.orangehrmlive.com/")
# driver.maximize_window()
# driver.find_element(By.XPATH,"//input[@id='txtUsername']").send_keys('admin')
# driver.find_element(By.XPATH,"//input[@id='txtPassword']").send_keys('admin123')
# driver.find_element(By.XPATH,"//input[@id='btnLogin']").click()
# driver.save_screenshot(os.getcwd()+"\\orangehrmlogin.png")
#

#headless mode
# from selenium import webdriver
# def headless_chrome():
#     from selenium.webdriver.chrome.service import Service
#     serv_obj=Service("C:\Program Files\chromedriver_win32\chromedriver")
#     ops = webdriver.ChromeOptions()
#     ops.headless = True
#     driver= webdriver.Chrome(service=serv_obj)
#     return driver
# driver = headless_chrome()
# driver.get("//opensource-demo.orangehrmlive.com/")
# print(driver.title)
# driver.close()
#
from selenium import webdriver
from selenium.webdriver.common.by import By
# import os
# location=os.getcwd()
# def Chrome_setup():
#     from selenium.webdriver.chrome.service import Service
#     serv_obj=Service("C:\Program Files\chromedriver_win32\chromedriver")
#     preference={'download.default_directory':location}
#     ops=webdriver.ChromeOptions()
#     ops.add_experimental_option('prefs',preference)
#     driver=webdriver.Chrome(service=serv_obj)
#     return driver
# driver= Chrome_setup()
# driver.get("https://opensource-demo.orangehrmlive.com/")
# driver.maximize_window()
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

#file--WORKBOOK--SHEET--ROW--CELL
# file=""
# workbook=openpyxl.load_workbook(file)
# sheet=workbook['sheet1']
# rows=sheet.max_row
# cols=sheet.max_coll
# #read all data from excel
# for r in range(2,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value,end=' ')
#         print()
#
#
# #write data into excel
# file=""
# workbook=openpyxl.load_workbook(file)
# sheet=workbook['data']
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value ='selenium'
# workbook.save(file)
# from excel import utilies
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import Select
# import openpyxl
# serv_obj =Service("")
# driver=webdriver.Chrome(service=serv_obj)
# file=("")
#
# rows=utilies.getrowcount(file,sheet)
# for r in range(2,rows+1):
#
#     pref=utilies.readdata(file,'sheet',r,1)
#     rateofint=utilies.readdata(file,'sheet',r,2)
#     per1=utilies.readdata(file,'sheet1',r,3)
#     per2=utilies.readdata(file,'sheet1',r,4)
#     freq=utilies.readdata(file,'sheet1',r,5)
#     exp_value=utilies.readdata(file,'sheet1',r,6)
#
#     driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(pref)
#     driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(rateofintrst)
#     driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(per1)
#     perioddrp = Select(driver.find_element(By.XPATH, "//select[@id='tenurePeriod']"))
#     perioddrp.select_by_visible_text(per2)
#     frequency = Select(driver.find_element(By.XPATH, "//select[@id='frequency']"))
#     frequency.select_by_visible_text(fre)
#     driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img").click()
#     act_value = driver.find_element(By.XPATH, "//span[@id='resp_matval']/strong").text
#
#     if float(exp_value) ==float(act_value):
#         print('test passed')
#         utilies.writedata(file,'sheet1',r,8,'passed')
#         utilies.fillgreencolor(file,'sheet1',r,8)
#     else:
#         print('test failed')
#         utilies.writedata(file,'sheet1',r,8,'failed')
#         utilies.fillredcolour(file,'sheet1',r,8)
#     driver.find_element(By.XPATH,"").click()
# driver.close()
#


# import mysql.connector
# try:
#     con=mysql.connector.connect(host='localhost',port=3306,username='root',password='root',database='mydb')
#     curs=con.cursor()
#     curs.execute('insert into student set kim=104')
#     con.commit()
#     con.close()
# except:
#     print('unsucessful')
# print('finished')
# import mysql.connector
# try:
#     con=mysql.connector.connect(host='ocalhost',port=3306,username='root',password='root',database='mydb')
#     curs=con.cursor()
#     curs.execute('delete from set mary=104')
#     con.commit()
#     con.close()
# except:
#     print('unsucessful')
# print('finished')
#
# import mysql.connector
# try:
#     con=mysql.connector.connect(host='myhost',port=3306,username='root',password='root',database='mydb')
#     curs=con.cursor()
#     curs.execute('update data in kim=104')
#     con.commit()
#     con.close
# except:
#     print('unsucessful')
# print('finished')
# import mysql.connector
# try:
#
#     con=mysql.connector.connect(host='localhost',port=3306,username='root',password='root'database='mydb')
#     curs=con.cursor()
#     curs.execute('insert into the student kim=104')
#     con.commit()
#     con.close()
# except:
#     print('unsucessful')
# print('finished')

# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.support.ui import Select
# import mysql.connector
#
# serv_obj=Service("")
# driver=webdriver.Chrome(service=serv_obj)
# con=mysql.connector.connect(host='localhost',port=3306,username='root',password='root',database='mydb')
# curs=con.cursor()
# curs.execute('select * from data')
#
# try:
#     for row in curs:
#         pref=row[0]
#         rateofint=row[1]
#         per1=row[2]
#         per2=row[3]
#         freq=row[4]
#         exp_value=row[5]
#
#         driver.find_element(By.XPATH,"").send_keys(pref)
#         driver.find_element(By.XPATH,"").send_keys(rateofint)
#         driver.find_element(By.XPATH,"").send_keys(per1)
#         drpdown=Select(driver.find_element(By.XPATH,""))
#         drpdown.select_by_visible_text(per2)
#         dropdwn=Select(driver.find_element(By.XPATH,""))
#         dropdwn.select_by_visible_text(freq)
#         act_value=driver.find_element(By.XPATH,"").text
#
#         if float(act_value) == float(exp_value):
#             print('test passed')
#         else:
#             print('failed')
#             driver.find_element(By.XPATH,"").click()
#         con.close()
# except:
#     print('unsucessful')
# driver.close()
#



from excel import utilies
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl
from openpyxl.styles import PatternFill

serv_obj=Service("")
driver=webdriver.Chrome(service=serv_obj)
driver.get("")
driver.maximize_window()
file=("")
noofrows=utilies.getrowcount(file,'data')

for r in range(2,noofrows+1):
    pref=utilies.readdata(file,'data',r,1)
    rateofintst=utilies.readdata(file,'data',r,2)
    per1=utilies.readdata(file,'data',r,3)
    per2=utilies.readdata(file,'data',r,4)
    freq=utilies.readdata(file,'data',r,5)
    exp_value=utilies.readdata(file,'data',r,6)

    driver.find_element(By.XPATH,"").send_keys(pref)
    driver.find_element(By.XPATH,"").send_keys(rateofintst)
    driver.find_element(By.XPATH,"").send_keys(per1)
    perioddrp=Select(driver.find_element(By.XPATH,""))
    perioddrp.select_by_visible_text(per2)
    frequencydrp=Select(driver.find_element(By.XPATH,""))
    frequencydrp.select_by_visible_text(freq)
    actvalue=driver.find_element(By.XPATH,"").text

    if float(exp_value) == float(actvalue):
        print('test passed')
        utilies.writedata(file,'data',r,8,'passed')
        utilies.fillgreencolor(file,'data',r,8)
    else:
        print('test failed')
        utilies.writedata(file,'data',r,8,'failed')
        utilies.fillredcolour(file,'data',r,8,)
    driver.find_element(By.XPATH,"").click()
driver.close()
import openpyxl
file=''
workbook=openpyxl.load_workbook(file)
sheet=workbook['sheet1']
rows=sheet.max_row
column=sheet.max_column
#read data
for r in range(1,rows+1):
    for c in range(1,column+1)
        print()

#write data of same value
import openpyxl
file=""
workbook=openpyxl.load_workbook(file)
sheet=workbook['data']
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value='selenium'
workbook.save(file)

#for diff data
file=''
workbook=openpyxl.load_workbook(file)
sheet=workbook['data']
sheet.cell(1,1).value='1'
sheet.cell(1,2).value='bharath'
sheet.cell(1,3).value='engineer'

sheet.cell(2,1).value='2'
sheet.cell(2,2).value='bharath'
sheet.cell(2,3).value='engineer'

sheet.cell(3,1).value='3'
sheet.cell(3,2).value='bharath'
sheet.cell(3,3).value='engineer'
workbook.save(file)

#read data from file
file=""
workbook=openpyxl.load_workbook(file)
sheet=workbook['sheet1']
rows=sheet.max_row
columns=sheet.max_column

for r in range(1,rows+1):
    for c in range(2,columns+1):
        print()

#write data
file=''
workbook=openpyxl.load_workbook(file)
sheet=workbook['data']
#for same data
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value='helloo'
workbook.save(file)
#for dif value
sheet.cell(1,1).value='1'
sheet.cell(1,2).value='bahrath'
sheet.cell(1,3).value='engineer'
workbook.save(file)




import mysql.connector


